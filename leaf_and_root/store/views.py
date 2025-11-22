import requests
import random
from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import TemplateView
from django.contrib.auth.mixins import UserPassesTestMixin
from django.contrib.auth.decorators import user_passes_test
from ordenes.models import Order, OrderItem
from catalogo.models import Product, Review, ProductClick, ProductSearchLog
from django.db.models import Q, Count, Sum
from django.http import HttpResponse
from openpyxl import Workbook
import datetime


# ---------- Solo admins ----------
def is_admin(user):
    return user.is_authenticated and user.is_staff

@user_passes_test(is_admin)
def admin_dashboard(request):
    products = Product.objects.all()
    return render(request, "admin/dashboard.html", {"products": products})

def redirect_to_home(request): 
    return redirect('product_list')

# ---------- Administración ----------
class AdminDashboardView(UserPassesTestMixin, TemplateView):
    template_name = "admin/dashboard.html"

    def test_func(self):
        return self.request.user.is_staff

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["product_count"] = Product.objects.count()
        context["order_count"] = Order.objects.count()
        context["user_count"] = Review.objects.values("customer").distinct().count()  # aproximado
        context["pending_reviews"] = Review.objects.filter(approved=False)

        # Top clics productos
        top_clicks_qs = (
            ProductClick.objects.values("product__name")
            .annotate(total=Count("id"))
            .order_by("-total")[:10]
        )
        context["top_clicks"] = list(top_clicks_qs)

        # Top búsquedas
        top_searches_qs = (
            ProductSearchLog.objects.values("query")
            .annotate(total=Count("id"))
            .order_by("-total")[:10]
        )
        context["top_searches"] = list(top_searches_qs)

        # Top comprados
        top_purchased_qs = (
            OrderItem.objects.values("product__name")
            .annotate(quantity=Sum("quantity"))
            .order_by("-quantity")[:10]
        )
        context["top_purchased"] = list(top_purchased_qs)

        # Stock productos (para gráfica)
        stock_qs = Product.objects.values("name", "stock").order_by("name")
        context["stock_data"] = list(stock_qs)

        # Productos por categoría
        category_qs = (
            Product.objects.values("category")
            .annotate(count=Count("id_product"))
            .order_by("-count")
        )
        context["category_counts"] = [c for c in category_qs if c.get("category")]

        # Preparar datos para Chart.js
        context["chart_clicks_labels"] = [x["product__name"] for x in top_clicks_qs]
        context["chart_clicks_values"] = [x["total"] for x in top_clicks_qs]
        context["chart_purchased_labels"] = [x["product__name"] for x in top_purchased_qs]
        context["chart_purchased_values"] = [x["quantity"] for x in top_purchased_qs]
        context["chart_category_labels"] = [x["category"] for x in context["category_counts"]]
        context["chart_category_values"] = [x["count"] for x in context["category_counts"]]
        context["chart_stock_labels"] = [x["name"] for x in stock_qs]
        context["chart_stock_values"] = [x["stock"] for x in stock_qs]
        return context

def export_dashboard_excel(request):
    if not request.user.is_authenticated or not request.user.is_staff:
        return HttpResponse(status=403)
    
    wb = Workbook()
    ws_clicks = wb.active
    ws_clicks.title = "Clicks" #type: ignore
    ws_clicks.append(["Producto", "Total Clics"]) #type: ignore
    for row in ProductClick.objects.values("product__name").annotate(total=Count("id")).order_by("-total"):
        ws_clicks.append([row["product__name"], row["total"]]) #type: ignore 

    ws_search = wb.create_sheet("Busquedas")
    ws_search.append(["Query", "Total", "Resultados promedio"])
    for row in ProductSearchLog.objects.values("query").annotate(total=Count("id")).order_by("-total"):
        avg_results = ProductSearchLog.objects.filter(query=row["query"]).aggregate(avg=Sum("results_count")/Count("id"))
        ws_search.append([row["query"], row["total"], int(avg_results.get("avg") or 0)])

    ws_purchased = wb.create_sheet("Comprados")
    ws_purchased.append(["Producto", "Cantidad"])
    for row in OrderItem.objects.values("product__name").annotate(quantity=Sum("quantity")).order_by("-quantity"):
        ws_purchased.append([row["product__name"], row["quantity"]])

    ws_stock = wb.create_sheet("Stock")
    ws_stock.append(["Producto", "Stock"])
    for p in Product.objects.values("name", "stock"):
        ws_stock.append([p["name"], p["stock"]])

    ws_cat = wb.create_sheet("Categorias")
    ws_cat.append(["Categoria", "Cantidad Productos"])
    for c in Product.objects.values("category").annotate(count=Count("id_product")).order_by("-count"):
        if c["category"]:
            ws_cat.append([c["category"], c["count"]])

    # Preparar respuesta
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"dashboard_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response["Content-Disposition"] = f"attachment; filename={filename}"
    wb.save(response)
    return response

def veg_recipe(request):
    # TheMealDB vegetarian/vegan filter endpoints
    veg_urls = [
        'https://www.themealdb.com/api/json/v1/1/filter.php?c=Vegetarian',
        'https://www.themealdb.com/api/json/v1/1/filter.php?c=Vegan',
    ]
    # Pick one category randomly
    url = random.choice(veg_urls)
    resp = requests.get(url, timeout=8)
    meals = resp.json().get('meals', [])
    recipe = None
    if meals:
        meal = random.choice(meals)
        meal_id = meal['idMeal']
        detail_url = f'https://www.themealdb.com/api/json/v1/1/lookup.php?i={meal_id}'
        detail_resp = requests.get(detail_url, timeout=8)
        recipe = detail_resp.json().get('meals', [{}])[0]
        # Prefer Spanish if available and language is es
        lang = getattr(request, 'LANGUAGE_CODE', 'en')
        if lang.startswith('es'):
            if recipe.get('strInstructionsES'):
                recipe['strInstructions'] = recipe['strInstructionsES']
                if recipe.get('strMealES'):
                    recipe['strMeal'] = recipe['strMealES']
    return render(request, 'veg_recipe.html', {'recipe': recipe})

